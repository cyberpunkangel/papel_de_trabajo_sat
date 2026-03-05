"""Generador de reportes Excel para los documentos procesados."""

from __future__ import annotations

import logging
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string

from ..models.documento import Documento

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Construye el archivo Excel con las hojas de detalle e integración."""

    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(color='FFFFFF', bold=True)
    TITLE_FONT = Font(size=14, bold=True)
    SUMMARY_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    MONETARY_KEYWORDS = (
        'INGRESO',
        'EXENTO',
        'PERDIDA',
        'ISR',
        'MONTO',
        'BASE',
        'INTERESES',
        'DIVIDENDOS',
    )
    BORDER_SIDE = Side(style='thin', color='FF4A4A4A')
    TABLE_BORDER = Border(
        left=BORDER_SIDE,
        right=BORDER_SIDE,
        top=BORDER_SIDE,
        bottom=BORDER_SIDE,
    )
    CURRENCY_FORMAT = '"$"#,##0.00'
    AMOUNT_COLUMN = 'B'
    TABULADOR_MATCH_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    COLUMNAS_NOMINAS = [
        'No',
        'TIPO DE INGRESO',
        'NOMBRE',
        'RFC',
        'INGRESO GRAVADO',
        'EXENTOS',
        'ISR RETENIDO',
        'UUID',
    ]
    COLUMNAS_INTERESES = [
        'No',
        'TIPO DE INGRESO',
        'NOMBRE',
        'RFC',
        'INTERESES NOMINALES',
        'INTERESES REALES',
        'PERDIDA',
        'ISR RETENIDO',
        'UUID',
    ]
    COLUMNAS_DIVIDENDOS = [
        'No',
        'TIPO DE INGRESO',
        'NOMBRE',
        'RFC',
        'MONTO DIVIDENDOS NACIONALES',
        'MONTO DIVIDENDOS EXTRANJEROS',
        'ISR ACREDITABLE MÉXICO',
        'ISR ACREDITABLE EXTRANJERO',
        'BASE DE LA RETENCIÓN',
        'MONTO RETENCIÓN',
        'UUID',
    ]
    COLUMNAS_ENAJENACION = [
        'No',
        'TIPO DE INGRESO',
        'NOMBRE',
        'RFC',
        'INGRESO GRAVADO',
        'EXENTOS',
        'PERDIDA',
        'ISR RETENIDO',
        'UUID',
    ]
    COLUMNAS_DEDUCCIONES = [
        'No',
        'TIPO DE INGRESO',
        'NOMBRE',
        'RFC',
        'USO CFDI',
        'MONTO DEDUCIBLE',
        'UUID',
    ]

    def __init__(self, ruta_salida: str, contribuyente: Optional[dict] = None, tabulador_isr: Optional[list[dict]] = None):
        self.ruta_salida = ruta_salida
        self.contribuyente = contribuyente or {}
        self._fila_intereses_reales_acumulables: Optional[int] = None
        self._fila_total_intereses_isr: Optional[int] = None
        self._fila_total_div_base: Optional[int] = None
        self._fila_total_div_isr: Optional[int] = None
        self._fila_enajenacion_integracion: Optional[int] = None
        self._fila_total_nominas: Optional[int] = None
        self._fila_total_deducciones: Optional[int] = None
        self.tabulador_isr = tabulador_isr or []
        self._tabulador_bounds: Optional[dict[str, int]] = None

    def crear_reporte(self, documentos: List[Documento]) -> None:
        if not documentos:
            raise ValueError('No hay documentos para generar el reporte')

        df = pd.DataFrame([doc.to_dict() for doc in documentos])
        if df.empty:
            raise ValueError('No fue posible construir la tabla con los documentos procesados')

        if 'UUID' in df.columns:
            df = df.drop_duplicates(subset=['UUID'], keep='first')

        df_nominas = df[df['TIPO DE INGRESO'] == 'nomina12'].copy()
        df_deducciones = df[df['TIPO DE INGRESO'] == 'deducciones_personales'].copy()
        df_dividendos = df[df['TIPO DE INGRESO'] == 'dividendos'].copy()
        df_enajenacion = df[df['TIPO DE INGRESO'] == 'enajenaciondeacciones'].copy()
        df_intereses = df[
            ~df['TIPO DE INGRESO'].isin({
                'nomina12',
                'deducciones_personales',
                'dividendos',
                'enajenaciondeacciones',
            })
        ].copy()

        logger.info(
            'Generando reporte Excel en %s (nominas=%s, deducciones=%s, intereses=%s, dividendos=%s, enajenacion=%s)',
            self.ruta_salida,
            len(df_nominas),
            len(df_deducciones),
            len(df_intereses),
            len(df_dividendos),
            len(df_enajenacion),
        )

        with pd.ExcelWriter(self.ruta_salida, engine='openpyxl') as writer:
            workbook = writer.book

            if not df_nominas.empty:
                self._crear_hoja(workbook, df_nominas, 'Nóminas', self.COLUMNAS_NOMINAS)
            if not df_deducciones.empty:
                self._crear_hoja(workbook, df_deducciones, 'Deducciones', self.COLUMNAS_DEDUCCIONES)
            if not df_intereses.empty:
                self._crear_hoja(workbook, df_intereses, 'Intereses', self.COLUMNAS_INTERESES)
            if not df_dividendos.empty:
                self._crear_hoja(workbook, df_dividendos, 'Dividendos', self.COLUMNAS_DIVIDENDOS)
            if not df_enajenacion.empty:
                self._crear_hoja(workbook, df_enajenacion, 'Enajenación de Acciones', self.COLUMNAS_ENAJENACION)

            self._crear_hoja_integracion(
                workbook,
                df_nominas,
                df_deducciones,
                df_intereses,
                df_dividendos,
                df_enajenacion,
            )
            self._crear_hoja_determinacion(workbook)

            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

            if workbook.sheetnames:
                workbook.active = workbook[workbook.sheetnames[0]]

    def _crear_hoja(
        self,
        workbook: Workbook,
        df: pd.DataFrame,
        nombre_hoja: str,
        columnas: List[str],
    ) -> None:
        df = df.copy()
        df = df.reset_index(drop=True)
        df.insert(0, 'No', range(1, len(df) + 1))

        for columna in columnas:
            if columna == 'No':
                continue
            if columna not in df.columns:
                df[columna] = 0.0 if self._es_columna_monetaria(columna) else ''

        df = df[columnas]

        ws = workbook.create_sheet(nombre_hoja)
        ws.append([nombre_hoja.upper()])
        ws['A1'].font = self.TITLE_FONT
        ws.append([])
        ws.append(columnas)

        for cell in ws[3]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        for _, row in df.iterrows():
            valores = []
            for columna in columnas:
                valor = row[columna]
                if pd.isna(valor):
                    valor = ''
                elif isinstance(valor, (int, float)):
                    valor = float(valor)
                valores.append(valor)
            ws.append(valores)

        for idx, columna in enumerate(columnas, start=1):
            letter = get_column_letter(idx)
            valores = [str(columna)]
            if columna in df.columns:
                valores.extend(df[columna].fillna('').astype(str).tolist())
            max_length = min(max(len(valor) for valor in valores) + 2, 45)
            ws.column_dimensions[letter].width = max_length

            if columna != 'No' and self._es_columna_monetaria(columna):
                for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=idx, max_col=idx):
                    row[0].number_format = '"$"#,##0.00'

    def _crear_hoja_integracion(
        self,
        workbook: Workbook,
        df_nominas: pd.DataFrame,
        df_deducciones: pd.DataFrame,
        df_intereses: pd.DataFrame,
        df_dividendos: pd.DataFrame,
        df_enajenacion: pd.DataFrame,
    ) -> None:
        ws = workbook.create_sheet('INTEGRACION', 0)
        periodo = self.contribuyente.get('periodo', '')
        ws['A1'] = 'NOMBRE:'
        ws['B1'] = self.contribuyente.get('nombre', '')
        ws['A2'] = 'RFC:'
        ws['B2'] = self.contribuyente.get('rfc', '')
        ws['A3'] = 'CURP:'
        ws['B3'] = self.contribuyente.get('curp', '')
        ws['A4'] = 'PERIODO:'
        ws['B4'] = periodo

        fila_titulo_nominas = 6
        ws[f'A{fila_titulo_nominas}'] = f"Nóminas {periodo}".strip()
        ws[f'A{fila_titulo_nominas}'].font = Font(bold=True)

        fila_header_nominas = fila_titulo_nominas + 1
        ws[f'B{fila_header_nominas}'] = 'TIPO DE INGRESO'
        ws[f'C{fila_header_nominas}'] = 'Nombre'
        ws[f'D{fila_header_nominas}'] = 'RFC'
        ws[f'E{fila_header_nominas}'] = 'INGRESOS ACUMULABLES'
        ws[f'F{fila_header_nominas}'] = 'INGRESOS EXENTOS'
        ws[f'G{fila_header_nominas}'] = 'IMPUESTO RETENIDO'
        self._style_header_row(ws, fila_header_nominas, 'B', 'G')

        fila_nominas_total = fila_header_nominas + 1
        ws[f'B{fila_nominas_total}'] = 'nomina12'
        ws[f'C{fila_nominas_total}'] = 'Total'
        ws[f'D{fila_nominas_total}'] = ''
        has_nominas = 'Nóminas' in workbook.sheetnames
        nomina_cols = {
            'tipo': self._column_letter(self.COLUMNAS_NOMINAS, 'TIPO DE INGRESO'),
            'rfc': self._column_letter(self.COLUMNAS_NOMINAS, 'RFC'),
            'gravado': self._column_letter(self.COLUMNAS_NOMINAS, 'INGRESO GRAVADO'),
            'exento': self._column_letter(self.COLUMNAS_NOMINAS, 'EXENTOS'),
            'isr': self._column_letter(self.COLUMNAS_NOMINAS, 'ISR RETENIDO'),
        }
        if has_nominas:
            ws[f'E{fila_nominas_total}'] = self._sum_column_formula(
                workbook, 'Nóminas', nomina_cols['gravado']
            )
            ws[f'F{fila_nominas_total}'] = self._sum_column_formula(
                workbook, 'Nóminas', nomina_cols['exento']
            )
            ws[f'G{fila_nominas_total}'] = self._sum_column_formula(
                workbook, 'Nóminas', nomina_cols['isr']
            )
        else:
            ws[f'E{fila_nominas_total}'] = 0
            ws[f'F{fila_nominas_total}'] = 0
            ws[f'G{fila_nominas_total}'] = 0

        self._fila_total_nominas = fila_nominas_total
        self._apply_table_border(ws, fila_header_nominas, fila_nominas_total, 'B', 'G')
        self._apply_currency_format(ws, ('E', 'F', 'G'), fila_nominas_total, fila_nominas_total)

        fila_cursor = fila_nominas_total + 1
        detalle_nominas = self._unique_instituciones(df_nominas)

        if detalle_nominas:
            ws[f'A{fila_cursor}'] = 'Detalle por patrón'
            ws[f'A{fila_cursor}'].font = Font(bold=True)
            fila_cursor += 1
            ws[f'B{fila_cursor}'] = 'TIPO DE INGRESO'
            ws[f'C{fila_cursor}'] = 'Nombre'
            ws[f'D{fila_cursor}'] = 'RFC'
            ws[f'E{fila_cursor}'] = 'INGRESOS ACUMULABLES'
            ws[f'F{fila_cursor}'] = 'INGRESOS EXENTOS'
            ws[f'G{fila_cursor}'] = 'IMPUESTO RETENIDO'
            fila_header_detalle = fila_cursor
            self._style_header_row(ws, fila_header_detalle, 'B', 'G')

            fila_cursor += 1
            fila_detalle_inicio = fila_cursor
            for tipo, nombre, rfc in detalle_nominas:
                ws[f'B{fila_cursor}'] = tipo
                ws[f'C{fila_cursor}'] = nombre
                ws[f'D{fila_cursor}'] = rfc
                if has_nominas:
                    ws[f'E{fila_cursor}'] = self._sumifs_formula(
                        'Nóminas',
                        nomina_cols['gravado'],
                        fila_cursor,
                        nomina_cols['tipo'],
                        nomina_cols['rfc'],
                    )
                    ws[f'F{fila_cursor}'] = self._sumifs_formula(
                        'Nóminas',
                        nomina_cols['exento'],
                        fila_cursor,
                        nomina_cols['tipo'],
                        nomina_cols['rfc'],
                    )
                    ws[f'G{fila_cursor}'] = self._sumifs_formula(
                        'Nóminas',
                        nomina_cols['isr'],
                        fila_cursor,
                        nomina_cols['tipo'],
                        nomina_cols['rfc'],
                    )
                else:
                    ws[f'E{fila_cursor}'] = 0
                    ws[f'F{fila_cursor}'] = 0
                    ws[f'G{fila_cursor}'] = 0
                fila_cursor += 1
            fila_detalle_fin = fila_cursor - 1
            self._apply_table_border(ws, fila_header_detalle, fila_detalle_fin, 'B', 'G')
            self._apply_currency_format(ws, ('E', 'F', 'G'), fila_detalle_inicio, fila_detalle_fin)
            fila_cursor += 1
        else:
            fila_cursor += 1

        fila_deducciones = max(fila_cursor, fila_nominas_total + 2)
        ws[f'A{fila_deducciones}'] = f"Deducciones personales {periodo}".strip()
        ws[f'A{fila_deducciones}'].font = Font(bold=True)

        fila_header_ded = fila_deducciones + 1
        ws[f'B{fila_header_ded}'] = 'TIPO DE INGRESO'
        ws[f'C{fila_header_ded}'] = 'Nombre'
        ws[f'D{fila_header_ded}'] = 'RFC'
        ws[f'E{fila_header_ded}'] = 'MONTO DEDUCIBLE'
        self._style_header_row(ws, fila_header_ded, 'B', 'E')

        fila_total_ded = fila_header_ded + 1
        ws[f'B{fila_total_ded}'] = 'deducciones_personales'
        ws[f'C{fila_total_ded}'] = 'Total'
        ws[f'D{fila_total_ded}'] = ''
        has_deducciones = 'Deducciones' in workbook.sheetnames
        ded_cols = {
            'tipo': self._column_letter(self.COLUMNAS_DEDUCCIONES, 'TIPO DE INGRESO'),
            'rfc': self._column_letter(self.COLUMNAS_DEDUCCIONES, 'RFC'),
            'monto': self._column_letter(self.COLUMNAS_DEDUCCIONES, 'MONTO DEDUCIBLE'),
        }
        if has_deducciones:
            ws[f'E{fila_total_ded}'] = self._sum_column_formula(
                workbook,
                'Deducciones',
                ded_cols['monto'],
            )
        else:
            ws[f'E{fila_total_ded}'] = 0

        self._fila_total_deducciones = fila_total_ded
        self._apply_table_border(ws, fila_header_ded, fila_total_ded, 'B', 'E')
        self._apply_currency_format(ws, ('E',), fila_total_ded, fila_total_ded)

        fila_cursor = fila_total_ded + 1
        detalle_deducciones = self._unique_instituciones(df_deducciones)
        if detalle_deducciones:
            ws[f'A{fila_cursor}'] = 'Detalle por emisor'
            ws[f'A{fila_cursor}'].font = Font(bold=True)
            fila_cursor += 1
            ws[f'B{fila_cursor}'] = 'TIPO DE INGRESO'
            ws[f'C{fila_cursor}'] = 'Nombre'
            ws[f'D{fila_cursor}'] = 'RFC'
            ws[f'E{fila_cursor}'] = 'MONTO DEDUCIBLE'
            fila_header_ded_detalle = fila_cursor
            self._style_header_row(ws, fila_header_ded_detalle, 'B', 'E')

            fila_cursor += 1
            fila_ded_inicio = fila_cursor
            for tipo, nombre, rfc in detalle_deducciones:
                ws[f'B{fila_cursor}'] = tipo
                ws[f'C{fila_cursor}'] = nombre
                ws[f'D{fila_cursor}'] = rfc
                if has_deducciones:
                    ws[f'E{fila_cursor}'] = self._sumifs_formula(
                        'Deducciones',
                        ded_cols['monto'],
                        fila_cursor,
                        ded_cols['tipo'],
                        ded_cols['rfc'],
                    )
                else:
                    ws[f'E{fila_cursor}'] = 0
                fila_cursor += 1
            fila_ded_fin = fila_cursor - 1
            self._apply_table_border(ws, fila_header_ded_detalle, fila_ded_fin, 'B', 'E')
            self._apply_currency_format(ws, ('E',), fila_ded_inicio, fila_ded_fin)
            fila_cursor += 1
        else:
            fila_cursor += 1

        fila_intereses = max(fila_cursor, fila_nominas_total + 2)
        ws[f'A{fila_intereses}'] = f"Intereses {periodo}".strip()
        ws[f'A{fila_intereses}'].font = Font(bold=True)

        fila_nominal = fila_intereses + 1
        fila_real = fila_nominal + 1
        fila_perdida = fila_real + 1
        fila_isr = fila_perdida + 1
        fila_real_acum = fila_isr + 1

        ws[f'B{fila_nominal}'] = 'MONTO TOTAL DE INTERESES NOMINALES'
        ws[f'C{fila_nominal}'] = self._round_formula(
            self._sum_column_formula(
                workbook,
                'Intereses',
                self._column_letter(self.COLUMNAS_INTERESES, 'INTERESES NOMINALES'),
            )
        )

        ws[f'B{fila_real}'] = 'MONTO TOTAL DE INTERESES REALES'
        ws[f'C{fila_real}'] = self._round_formula(
            self._sum_column_formula(
                workbook,
                'Intereses',
                self._column_letter(self.COLUMNAS_INTERESES, 'INTERESES REALES'),
            )
        )

        ws[f'B{fila_perdida}'] = 'MONTO TOTAL DE PERDIDA'
        ws[f'C{fila_perdida}'] = self._round_formula(
            self._sum_column_formula(
                workbook,
                'Intereses',
                self._column_letter(self.COLUMNAS_INTERESES, 'PERDIDA'),
            )
        )

        ws[f'B{fila_isr}'] = 'MONTO TOTAL DEL ISR RETENIDO'
        ws[f'C{fila_isr}'] = self._round_formula(
            self._sum_column_formula(
                workbook,
                'Intereses',
                self._column_letter(self.COLUMNAS_INTERESES, 'ISR RETENIDO'),
            )
        )
        self._fila_total_intereses_isr = fila_isr

        ws[f'B{fila_real_acum}'] = 'INTERESES REALES ACUMULABLES'
        ws[f'C{fila_real_acum}'] = f"=ROUND(C{fila_real}-C{fila_perdida},0)"
        self._fila_intereses_reales_acumulables = fila_real_acum

        self._apply_table_border(ws, fila_nominal, fila_real_acum, 'B', 'C')
        self._apply_currency_format(ws, ('C',), fila_nominal, fila_real_acum)

        fila_cursor = fila_real_acum + 2
        ws[f'B{fila_cursor}'] = 'TIPO DE INGRESO'
        ws[f'C{fila_cursor}'] = 'Nombre'
        ws[f'D{fila_cursor}'] = 'RFC'
        ws[f'E{fila_cursor}'] = 'INTERESES NOMINALES'
        ws[f'F{fila_cursor}'] = 'INTERESES REALES'
        ws[f'G{fila_cursor}'] = 'PERDIDA'
        ws[f'H{fila_cursor}'] = 'ISR RETENIDO'
        fila_header_intereses = fila_cursor
        self._style_header_row(ws, fila_header_intereses, 'B', 'H')

        fila_cursor += 1
        fila_interes_inicio = fila_cursor
        instituciones_intereses = self._unique_instituciones(df_intereses)
        has_intereses = 'Intereses' in workbook.sheetnames
        interes_cols = {
            'tipo': self._column_letter(self.COLUMNAS_INTERESES, 'TIPO DE INGRESO'),
            'rfc': self._column_letter(self.COLUMNAS_INTERESES, 'RFC'),
            'nominal': self._column_letter(self.COLUMNAS_INTERESES, 'INTERESES NOMINALES'),
            'real': self._column_letter(self.COLUMNAS_INTERESES, 'INTERESES REALES'),
            'perdida': self._column_letter(self.COLUMNAS_INTERESES, 'PERDIDA'),
            'isr': self._column_letter(self.COLUMNAS_INTERESES, 'ISR RETENIDO'),
        }

        for tipo, nombre, rfc in instituciones_intereses:
            ws[f'B{fila_cursor}'] = tipo
            ws[f'C{fila_cursor}'] = nombre
            ws[f'D{fila_cursor}'] = rfc
            if has_intereses:
                ws[f'E{fila_cursor}'] = self._sumifs_formula(
                    'Intereses',
                    interes_cols['nominal'],
                    fila_cursor,
                    interes_cols['tipo'],
                    interes_cols['rfc'],
                )
                ws[f'F{fila_cursor}'] = self._sumifs_formula(
                    'Intereses',
                    interes_cols['real'],
                    fila_cursor,
                    interes_cols['tipo'],
                    interes_cols['rfc'],
                )
                ws[f'G{fila_cursor}'] = self._sumifs_formula(
                    'Intereses',
                    interes_cols['perdida'],
                    fila_cursor,
                    interes_cols['tipo'],
                    interes_cols['rfc'],
                )
                ws[f'H{fila_cursor}'] = self._sumifs_formula(
                    'Intereses',
                    interes_cols['isr'],
                    fila_cursor,
                    interes_cols['tipo'],
                    interes_cols['rfc'],
                )
            else:
                ws[f'E{fila_cursor}'] = 0
                ws[f'F{fila_cursor}'] = 0
                ws[f'G{fila_cursor}'] = 0
                ws[f'H{fila_cursor}'] = 0
            fila_cursor += 1

        tiene_detalle_intereses = fila_cursor - 1 >= fila_interes_inicio
        fila_interes_fin = (fila_cursor - 1) if tiene_detalle_intereses else None

        if fila_interes_fin is not None:
            ws[f'C{fila_nominal}'] = self._round_formula(
                f"=SUM(E{fila_interes_inicio}:E{fila_interes_fin})"
            )
            ws[f'C{fila_real}'] = self._round_formula(
                f"=SUM(F{fila_interes_inicio}:F{fila_interes_fin})"
            )
            ws[f'C{fila_perdida}'] = self._round_formula(
                f"=SUM(G{fila_interes_inicio}:G{fila_interes_fin})"
            )
            ws[f'C{fila_isr}'] = self._round_formula(
                f"=SUM(H{fila_interes_inicio}:H{fila_interes_fin})"
            )

        borde_final = fila_interes_fin if fila_interes_fin is not None else fila_header_intereses
        self._apply_table_border(ws, fila_header_intereses, borde_final, 'B', 'H')
        if fila_interes_fin is not None:
            self._apply_currency_format(ws, ('E', 'F', 'G', 'H'), fila_interes_inicio, fila_interes_fin)

        fila_cursor += 1
        ws[f'A{fila_cursor}'] = f"Dividendos {periodo}".strip()
        ws[f'A{fila_cursor}'].font = Font(bold=True)

        fila_cursor += 1
        ws[f'B{fila_cursor}'] = 'DIVIDENDOS ACUMULABLES'
        self._fila_total_div_base = fila_cursor
        ws[f'C{fila_cursor}'] = self._sum_column_formula(
            workbook,
            'Dividendos',
            self._column_letter(self.COLUMNAS_DIVIDENDOS, 'BASE DE LA RETENCIÓN'),
        )

        fila_cursor += 1
        ws[f'B{fila_cursor}'] = 'MONTO TOTAL DEL ISR RETENIDO'
        fila_total_div_isr = fila_cursor
        ws[f'C{fila_cursor}'] = self._sum_column_formula(
            workbook,
            'Dividendos',
            self._column_letter(self.COLUMNAS_DIVIDENDOS, 'MONTO RETENCIÓN'),
        )
        self._fila_total_div_isr = fila_total_div_isr

        self._apply_table_border(ws, self._fila_total_div_base, fila_total_div_isr, 'B', 'C')
        self._apply_currency_format(ws, ('C',), self._fila_total_div_base, fila_total_div_isr)

        fila_cursor += 2
        ws[f'B{fila_cursor}'] = 'TIPO DE INGRESO'
        ws[f'C{fila_cursor}'] = 'Nombre'
        ws[f'D{fila_cursor}'] = 'RFC'
        ws[f'E{fila_cursor}'] = 'BASE DE LA RETENCIÓN'
        ws[f'F{fila_cursor}'] = 'ISR RETENIDO'
        fila_header_div_detalle = fila_cursor
        self._style_header_row(ws, fila_header_div_detalle, 'B', 'F')

        fila_cursor += 1
        primera_fila_div = fila_cursor

        instituciones_dividendos = self._unique_instituciones(df_dividendos)
        has_dividendos = 'Dividendos' in workbook.sheetnames
        dividend_cols = {
            'tipo': self._column_letter(self.COLUMNAS_DIVIDENDOS, 'TIPO DE INGRESO'),
            'rfc': self._column_letter(self.COLUMNAS_DIVIDENDOS, 'RFC'),
            'base': self._column_letter(self.COLUMNAS_DIVIDENDOS, 'BASE DE LA RETENCIÓN'),
            'isr': self._column_letter(self.COLUMNAS_DIVIDENDOS, 'MONTO RETENCIÓN'),
        }

        for tipo, nombre, rfc in instituciones_dividendos:
            ws[f'B{fila_cursor}'] = tipo
            ws[f'C{fila_cursor}'] = nombre
            ws[f'D{fila_cursor}'] = rfc
            if has_dividendos:
                ws[f'E{fila_cursor}'] = self._sumifs_formula(
                    'Dividendos',
                    dividend_cols['base'],
                    fila_cursor,
                    dividend_cols['tipo'],
                    dividend_cols['rfc'],
                )
                ws[f'F{fila_cursor}'] = self._sumifs_formula(
                    'Dividendos',
                    dividend_cols['isr'],
                    fila_cursor,
                    dividend_cols['tipo'],
                    dividend_cols['rfc'],
                )
            else:
                ws[f'E{fila_cursor}'] = 0
                ws[f'F{fila_cursor}'] = 0
            fila_cursor += 1

        ultima_fila_div = max(fila_cursor - 1, primera_fila_div)
        if ultima_fila_div >= primera_fila_div:
            ws[f'C{self._fila_total_div_base}'] = f"=SUM(E{primera_fila_div}:E{ultima_fila_div})"
            ws[f'C{fila_total_div_isr}'] = f"=SUM(F{primera_fila_div}:F{ultima_fila_div})"

        self._apply_table_border(ws, fila_header_div_detalle, max(fila_header_div_detalle, ultima_fila_div), 'B', 'F')
        self._apply_currency_format(ws, ('E', 'F'), primera_fila_div, ultima_fila_div)

        fila_cursor += 1
        ws[f'A{fila_cursor}'] = f"Enajenación de Acciones {periodo}".strip()
        ws[f'A{fila_cursor}'].font = Font(bold=True)

        fila_cursor += 1
        ws[f'B{fila_cursor}'] = 'TIPO DE INGRESO'
        ws[f'C{fila_cursor}'] = 'Nombre'
        ws[f'D{fila_cursor}'] = 'RFC'
        ws[f'E{fila_cursor}'] = 'INGRESO GRAVADO'
        fila_header_enajenacion = fila_cursor
        self._style_header_row(ws, fila_header_enajenacion, 'B', 'E')

        fila_cursor += 1
        enajenacion_inicio = fila_cursor
        instituciones_enajenacion = self._unique_instituciones(df_enajenacion)
        has_enajenacion = 'Enajenación de Acciones' in workbook.sheetnames
        enajenacion_cols = {
            'tipo': self._column_letter(self.COLUMNAS_ENAJENACION, 'TIPO DE INGRESO'),
            'rfc': self._column_letter(self.COLUMNAS_ENAJENACION, 'RFC'),
            'ingreso': self._column_letter(self.COLUMNAS_ENAJENACION, 'INGRESO GRAVADO'),
        }

        for tipo, nombre, rfc in instituciones_enajenacion:
            ws[f'B{fila_cursor}'] = tipo
            ws[f'C{fila_cursor}'] = nombre
            ws[f'D{fila_cursor}'] = rfc
            if has_enajenacion:
                ws[f'E{fila_cursor}'] = self._sumifs_formula(
                    'Enajenación de Acciones',
                    enajenacion_cols['ingreso'],
                    fila_cursor,
                    enajenacion_cols['tipo'],
                    enajenacion_cols['rfc'],
                )
            else:
                ws[f'E{fila_cursor}'] = 0
            fila_cursor += 1

        enajenacion_fin = fila_cursor - 1
        if instituciones_enajenacion:
            self._apply_table_border(ws, fila_header_enajenacion, enajenacion_fin, 'B', 'E')
            self._apply_currency_format(ws, ('E',), enajenacion_inicio, enajenacion_fin)
        self._fila_enajenacion_integracion = enajenacion_fin if instituciones_enajenacion else None

        for col in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'):
            ws.column_dimensions[col].width = 22 if col in {'A', 'B'} else 18

    def _crear_hoja_determinacion(self, workbook: Workbook) -> None:
        if 'INTEGRACION' not in workbook.sheetnames:
            return

        ws = workbook.create_sheet('DETERMINACION', 0)
        periodo = self.contribuyente.get('periodo', '')
        amount_col = self.AMOUNT_COLUMN
        amount_cell = lambda row: f'{amount_col}{row}'
        ws['A1'] = 'NOMBRE:'
        ws['B1'] = self.contribuyente.get('nombre', '')
        ws['A2'] = 'RFC:'
        ws['B2'] = self.contribuyente.get('rfc', '')
        ws['A3'] = 'CURP:'
        ws['B3'] = self.contribuyente.get('curp', '')
        ws['A4'] = 'PERIODO:'
        ws['B4'] = periodo

        ws['A6'] = 'CONCEPTO'
        ws[amount_cell(6)] = 'MONTO'
        ws['A6'].font = Font(bold=True, size=12)
        ws[amount_cell(6)].font = Font(bold=True, size=12)

        nomina_row = self._fila_total_nominas or 8
        row_base = 15
        row_limit = row_base + 1
        row_excess = row_limit + 1
        row_rate = row_excess + 1
        row_marginal = row_rate + 1
        row_fixed = row_marginal + 1
        row_total_isr = row_fixed + 1
        row_subsidio = row_total_isr + 1
        row_isr_neto = row_subsidio + 1
        row_retenciones = row_isr_neto + 1
        row_final = row_retenciones + 2  # deja una fila en blanco

        ws['A7'] = 'Ingresos por sueldos y salarios'
        if self._fila_total_nominas:
            ws[amount_cell(7)] = f"=INTEGRACION!E{nomina_row}+INTEGRACION!F{nomina_row}"
        else:
            ws[amount_cell(7)] = '=INTEGRACION!E8+INTEGRACION!F8'

        ws['A8'] = 'Más: Intereses reales acumulables'
        if self._fila_intereses_reales_acumulables:
            ws[amount_cell(8)] = f"=INTEGRACION!C{self._fila_intereses_reales_acumulables}"
        else:
            ws[amount_cell(8)] = 0

        ws['A9'] = 'Más: Dividendos acumulables'
        ws[amount_cell(9)] = f"=INTEGRACION!C{self._fila_total_div_base}" if self._fila_total_div_base else 0

        ws['A10'] = 'Más: Ganancias por enajenación de acciones'
        ws[amount_cell(10)] = f"=INTEGRACION!E{self._fila_enajenacion_integracion}" if self._fila_enajenacion_integracion else 0

        ws['A11'] = 'Ingresos acumulables por el Capítulo VI'
        ws[amount_cell(11)] = f"={amount_cell(8)}+{amount_cell(9)}+{amount_cell(10)}"
        ws['A11'].font = Font(bold=True)

        ws['A12'] = 'Total de ingresos acumulables'
        ws[amount_cell(12)] = f"={amount_cell(7)}+{amount_cell(11)}"
        ws['A12'].font = Font(bold=True, size=11)
        ws[amount_cell(12)].font = Font(bold=True)
        ws[amount_cell(12)].fill = self.SUMMARY_FILL

        ws['A13'] = 'EXENTOS'
        if self._fila_total_nominas:
            ws[amount_cell(13)] = f"=INTEGRACION!F{nomina_row}"
        else:
            ws[amount_cell(13)] = '=INTEGRACION!F8'

        raw_deducciones = self.contribuyente.get('deducciones_personales', 0)
        try:
            deducciones_ajuste = float(raw_deducciones)
        except (TypeError, ValueError):
            deducciones_ajuste = 0.0
        ws['A14'] = 'DEDUCCIONES PERSONALES'
        deducciones_auto_formula = (
            f"INTEGRACION!E{self._fila_total_deducciones}"
            if self._fila_total_deducciones
            else '0'
        )
        if deducciones_ajuste != 0:
            ws[amount_cell(14)] = f"=ROUND({deducciones_auto_formula}+{deducciones_ajuste},0)"
        else:
            ws[amount_cell(14)] = f"={deducciones_auto_formula}"

        ws['A15'] = 'Base gravable para ISR'
        ws[amount_cell(row_base)] = f"=ROUND({amount_cell(12)}-{amount_cell(13)}-{amount_cell(14)},0)"
        ws['A15'].font = Font(bold=True)
        ws[amount_cell(row_base)].font = Font(bold=True)

        self._write_tabulador_isr(ws, row_base, 'D')

        lower_range = self._tabulador_range(0)
        upper_range = self._tabulador_range(1)
        fixed_range = self._tabulador_range(2)
        rate_range = self._tabulador_range(3)

        ws[f'A{row_limit}'] = 'Límite inferior'
        if lower_range:
            ws[amount_cell(row_limit)] = f"=LOOKUP({amount_cell(row_base)},{lower_range})"
        else:
            ws[amount_cell(row_limit)] = 0

        ws[f'A{row_excess}'] = 'Excedente sobre límite inferior'
        ws[amount_cell(row_excess)] = f"=MAX(0,{amount_cell(row_base)}-{amount_cell(row_limit)})"

        ws[f'A{row_rate}'] = '% Mg'
        if lower_range and rate_range:
            ws[amount_cell(row_rate)] = (
                f"=IFERROR(INDEX({rate_range},MATCH({amount_cell(row_limit)},{lower_range},0)),0)"
            )
        else:
            ws[amount_cell(row_rate)] = 0

        ws[f'A{row_marginal}'] = 'Impuesto marginal'
        ws[amount_cell(row_marginal)] = f"=ROUND({amount_cell(row_excess)}*{amount_cell(row_rate)},2)"

        ws[f'A{row_fixed}'] = 'Cuota fija'
        if lower_range and fixed_range:
            ws[amount_cell(row_fixed)] = (
                f"=IFERROR(INDEX({fixed_range},MATCH({amount_cell(row_limit)},{lower_range},0)),0)"
            )
        else:
            ws[amount_cell(row_fixed)] = 0

        ws[f'A{row_total_isr}'] = 'Total ISR'
        ws[amount_cell(row_total_isr)] = f"=ROUND({amount_cell(row_marginal)}+{amount_cell(row_fixed)},0)"

        raw_subsidio = self.contribuyente.get('subsidio_empleo', 0)
        try:
            subsidio = float(raw_subsidio)
        except (TypeError, ValueError):
            subsidio = 0.0
        ws[f'A{row_subsidio}'] = 'Subsidio al empleo'
        ws[amount_cell(row_subsidio)] = subsidio

        ws[f'A{row_isr_neto}'] = 'ISR neto'
        ws[amount_cell(row_isr_neto)] = f"=ROUND({amount_cell(row_total_isr)}-{amount_cell(row_subsidio)},0)"

        ws[f'A{row_retenciones}'] = 'Menos: ISR retenido por sueldos y salarios'
        retention_terms: list[str] = []
        if self._fila_total_nominas:
            retention_terms.append(f"INTEGRACION!G{nomina_row}")
        else:
            retention_terms.append('INTEGRACION!G8')
        if self._fila_total_intereses_isr:
            retention_terms.append(f"INTEGRACION!C{self._fila_total_intereses_isr}")
        if self._fila_total_div_isr:
            retention_terms.append(f"INTEGRACION!C{self._fila_total_div_isr}")
        ws[amount_cell(row_retenciones)] = f"={'+'.join(retention_terms)}"

        ws[f'A{row_final}'] = 'A pagar (-a favor)'
        ws[amount_cell(row_final)] = f"=ROUND({amount_cell(row_isr_neto)}-{amount_cell(row_retenciones)},0)"
        ws[f'A{row_final}'].font = Font(bold=True)
        ws[amount_cell(row_final)].font = Font(bold=True)

        for fila in [7, 8, 9, 10, 11, 12, 13, 14, row_base, row_limit, row_excess,
                     row_marginal, row_fixed, row_total_isr, row_subsidio, row_isr_neto,
                     row_retenciones, row_final]:
            ws[amount_cell(fila)].number_format = '"$"#,##0.00'
        ws[amount_cell(row_rate)].number_format = '0.00%'

        ws.column_dimensions['A'].width = 55
        ws.column_dimensions[amount_col].width = 22
        for column in ('D', 'E', 'F', 'G'):
            ws.column_dimensions[column].width = 18

    def _style_header_row(self, ws, row: int, start_col: str, end_col: str) -> None:
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        for col in range(start_idx, end_idx + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _apply_table_border(self, ws, start_row: int, end_row: int, start_col: str, end_col: str) -> None:
        if end_row < start_row:
            return
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_idx, max_col=end_idx):
            for cell in row:
                cell.border = self.TABLE_BORDER

    def _apply_currency_format(self, ws, columns: tuple[str, ...], start_row: int, end_row: int) -> None:
        if end_row < start_row:
            return
        for column in columns:
            idx = column_index_from_string(column)
            for row in range(start_row, end_row + 1):
                cell = ws.cell(row=row, column=idx)
                cell.number_format = self.CURRENCY_FORMAT

    def _round_formula(self, formula: int | str, decimals: int = 0) -> int | str:
        if isinstance(formula, str) and formula.startswith('='):
            inner = formula[1:]
            return f"=ROUND({inner},{decimals})"
        return formula

    def _write_tabulador_isr(self, ws, start_row: int, start_column: str) -> None:
        if not self.tabulador_isr:
            self._tabulador_bounds = None
            return

        columns = [
            ('Límite inferior $', 'lower_limit', self.CURRENCY_FORMAT),
            ('Límite superior $', 'upper_limit', self.CURRENCY_FORMAT),
            ('Cuota fija $', 'fixed_fee', self.CURRENCY_FORMAT),
            ('% para aplicarse', 'rate', '0.00%'),
        ]

        start_idx = column_index_from_string(start_column)
        header_row = start_row

        for offset, (label, _, _) in enumerate(columns):
            cell = ws.cell(row=header_row, column=start_idx + offset, value=label)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row = header_row + 1
        for entry in self.tabulador_isr:
            for offset, (_, key, fmt) in enumerate(columns):
                value = entry.get(key)
                cell = ws.cell(row=current_row, column=start_idx + offset)
                is_number = isinstance(value, (int, float))
                cell.number_format = fmt if is_number else '@'
                if value in (None, ''):
                    cell.value = ''
                else:
                    cell.value = value
            current_row += 1

        end_column = get_column_letter(start_idx + len(columns) - 1)
        self._apply_table_border(ws, header_row, current_row - 1, start_column, end_column)
        for offset in range(len(columns)):
            column_letter = get_column_letter(start_idx + offset)
            ws.column_dimensions[column_letter].width = 18

        data_start = header_row + 1
        data_end = current_row - 1
        if data_end >= data_start:
            self._tabulador_bounds = {
                'start_col': start_idx,
                'data_start': data_start,
                'data_end': data_end,
            }
        else:
            self._tabulador_bounds = None
        if data_end >= data_start:
            lower_col_letter = start_column
            upper_col_letter = get_column_letter(start_idx + 1)
            data_range = f"{start_column}{data_start}:{end_column}{data_end}"
            lower_ref = f"${lower_col_letter}{data_start}"
            upper_ref = f"${upper_col_letter}{data_start}"
            amount_ref = f"${self.AMOUNT_COLUMN}${start_row}"
            formula = (
                f"AND({amount_ref}>={lower_ref},"
                f"OR({upper_ref}=\"\",{amount_ref}<={upper_ref}))"
            )
            ws.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=[formula],
                    stopIfTrue=False,
                    fill=self.TABULADOR_MATCH_FILL,
                ),
            )

    def _sum_column_formula(self, workbook: Workbook, sheet_name: str, column_letter: Optional[str]) -> int | str:
        if not column_letter or sheet_name not in workbook.sheetnames:
            return 0
        safe = sheet_name.replace("'", "''")
        return f"=SUM('{safe}'!${column_letter}:${column_letter})"

    def _sumifs_formula(
        self,
        sheet_name: str,
        target_column: Optional[str],
        row_number: int,
        tipo_column: Optional[str],
        rfc_column: Optional[str],
    ) -> int | str:
        if not all([target_column, tipo_column, rfc_column]):
            return 0
        safe = sheet_name.replace("'", "''")
        return (
            f"=ROUND(SUMIFS('{safe}'!${target_column}:${target_column},"
            f"'{safe}'!${rfc_column}:${rfc_column},INTEGRACION!D{row_number},"
            f"'{safe}'!${tipo_column}:${tipo_column},INTEGRACION!B{row_number}),0)"
        )

    def _column_letter(self, columnas: List[str], nombre: str) -> Optional[str]:
        if nombre not in columnas:
            return None
        idx = columnas.index(nombre) + 1
        return get_column_letter(idx)

    def _unique_instituciones(self, df: pd.DataFrame) -> List[tuple[str, str, str]]:
        if df is None or df.empty:
            return []
        instituciones: dict[str, tuple[str, str, str]] = {}
        for _, row in df.iterrows():
            rfc = (row.get('RFC') or '').strip()
            if not rfc:
                continue
            key = rfc.upper()
            if key not in instituciones:
                instituciones[key] = (
                    row.get('TIPO DE INGRESO', ''),
                    row.get('NOMBRE', ''),
                    rfc,
                )
        return sorted(instituciones.values(), key=lambda item: item[2])

    def _tabulador_range(self, offset: int) -> Optional[str]:
        if not self._tabulador_bounds:
            return None
        data_start = self._tabulador_bounds['data_start']
        data_end = self._tabulador_bounds['data_end']
        if data_end < data_start:
            return None
        column_letter = get_column_letter(self._tabulador_bounds['start_col'] + offset)
        return f"${column_letter}${data_start}:${column_letter}${data_end}"

    def _es_columna_monetaria(self, nombre: str) -> bool:
        upper_name = nombre.upper()
        return any(keyword in upper_name for keyword in self.MONETARY_KEYWORDS)

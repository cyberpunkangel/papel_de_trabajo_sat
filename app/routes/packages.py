"""
SAT package preparation routes.
Creates ZIP packages under 4 MB from downloaded XML files and workpapers.
"""

import os
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import quote

from fastapi import APIRouter, Form
from fastapi.responses import JSONResponse
from openpyxl import load_workbook

BASE_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', '..'))
DESCARGAS_DIR = os.path.join(BASE_DIR, 'descargas')
REPORTES_DIR = os.path.join(BASE_DIR, 'reportes')
PACKAGES_ROOT = os.path.join(REPORTES_DIR, 'paquetes_sat')

MAX_ZIP_BYTES = 4 * 1024 * 1024
SOURCE_FOLDERS = (
    'CONSTANCIAS_DE_RETENCIONES',
    'OTROS_CFDI',
    'RECIBOS_DE_NOMINA',
)

PACKAGE_NAME_BY_FOLDER = {
    'OTROS_CFDI': 'DEDUCCIONES_PERSONALES',
}

router = APIRouter()


def _normalize_period(value: Optional[str]) -> str:
    trimmed = (value or '').strip()
    digits = re.sub(r'[^0-9]', '', trimmed)
    return digits if len(digits) == 4 else ''


def _human_size(size_bytes: int) -> str:
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    val = float(size_bytes)
    idx = 0
    while val >= 1024 and idx < len(units) - 1:
        val /= 1024
        idx += 1
    return f'{val:.2f} {units[idx]}'


def _find_report_for_period(period: str) -> Optional[str]:
    if not os.path.isdir(REPORTES_DIR):
        return None

    candidates: List[str] = []
    for filename in os.listdir(REPORTES_DIR):
        full_path = os.path.join(REPORTES_DIR, filename)
        if not os.path.isfile(full_path):
            continue
        if not filename.lower().endswith('.xlsx'):
            continue
        if period not in filename:
            continue
        candidates.append(full_path)

    if not candidates:
        return None

    candidates.sort(key=lambda path: os.path.getmtime(path), reverse=True)
    return candidates[0]


def _collect_xml_entries(period: str, folder_name: str) -> List[Tuple[str, str]]:
    # Solo se consideran XML vigentes para generar paquetes SAT.
    vigentes_dir = os.path.join(DESCARGAS_DIR, period, folder_name, 'VIGENTES')
    if not os.path.isdir(vigentes_dir):
        return []

    entries: List[Tuple[str, str]] = []
    for dirpath, _dirnames, filenames in os.walk(vigentes_dir):
        for filename in filenames:
            if not filename.lower().endswith('.xml'):
                continue
            full_path = os.path.join(dirpath, filename)
            rel_path = os.path.relpath(full_path, vigentes_dir).replace(os.sep, '/')
            entries.append((full_path, rel_path))

    entries.sort(key=lambda item: item[1].lower())
    return entries


def _normalize_uuid(value: Optional[str]) -> str:
    return (value or '').strip().upper()


def _extract_deducciones_uuids(report_file: str) -> Set[str]:
    uuids: Set[str] = set()

    workbook = load_workbook(report_file, read_only=True, data_only=True)
    try:
        if 'Deducciones' not in workbook.sheetnames:
            return uuids

        sheet = workbook['Deducciones']
        uuid_col_idx: Optional[int] = None
        tipo_col_idx: Optional[int] = None
        header_row_idx: Optional[int] = None

        max_header_scan = min(sheet.max_row, 12)
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_header_scan, values_only=True),
            start=1,
        ):
            normalized_row = [str(value).strip().upper() if value is not None else '' for value in row]
            if 'UUID' in normalized_row:
                header_row_idx = row_idx
                uuid_col_idx = normalized_row.index('UUID')
                if 'TIPO DE INGRESO' in normalized_row:
                    tipo_col_idx = normalized_row.index('TIPO DE INGRESO')
                break

        if header_row_idx is None or uuid_col_idx is None:
            return uuids

        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            raw_uuid = row[uuid_col_idx] if len(row) > uuid_col_idx else None
            normalized_uuid = _normalize_uuid(str(raw_uuid) if raw_uuid is not None else '')
            if not normalized_uuid:
                continue

            if tipo_col_idx is not None and len(row) > tipo_col_idx:
                raw_tipo = row[tipo_col_idx]
                normalized_tipo = str(raw_tipo).strip().lower() if raw_tipo is not None else ''
                if normalized_tipo and normalized_tipo != 'deducciones_personales':
                    continue

            uuids.add(normalized_uuid)

        return uuids
    finally:
        workbook.close()


def _extract_xml_uuid(xml_path: str) -> str:
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        for elem in root.iter():
            if 'TimbreFiscalDigital' in elem.tag:
                return _normalize_uuid(elem.attrib.get('UUID'))
    except Exception:
        return ''
    return ''


def _filter_entries_by_allowed_uuids(
    entries: List[Tuple[str, str]],
    allowed_uuids: Set[str],
) -> List[Tuple[str, str]]:
    if not entries or not allowed_uuids:
        return []

    filtered: List[Tuple[str, str]] = []
    for source_path, relative_path in entries:
        xml_uuid = _extract_xml_uuid(source_path)
        if xml_uuid and xml_uuid in allowed_uuids:
            filtered.append((source_path, relative_path))
    return filtered


def _write_zip(zip_path: str, entries: List[Tuple[str, str]]) -> None:
    with zipfile.ZipFile(zip_path, mode='w', compression=zipfile.ZIP_DEFLATED) as zip_file:
        for source_path, arc_name in entries:
            zip_file.write(source_path, arc_name)


def _split_entries_by_zip_limit(
    entries: List[Tuple[str, str]],
    max_bytes: int,
    temp_dir: str,
    temp_prefix: str,
) -> List[List[Tuple[str, str]]]:
    if not entries:
        return []

    parts: List[List[Tuple[str, str]]] = []
    pending: List[Tuple[str, str]] = []

    for index, entry in enumerate(entries, start=1):
        trial = pending + [entry]
        temp_zip = os.path.join(temp_dir, f'.tmp_{temp_prefix}_{index}.zip')

        _write_zip(temp_zip, trial)
        trial_size = os.path.getsize(temp_zip)
        try:
            os.unlink(temp_zip)
        except OSError:
            pass

        if trial_size <= max_bytes:
            pending = trial
            continue

        if not pending:
            raise ValueError(
                f'El archivo {os.path.basename(entry[0])} supera el limite de 4 MB '
                'aun en un ZIP individual.'
            )

        parts.append(pending)
        pending = [entry]

    if pending:
        parts.append(pending)

    return parts


def _build_download_url(file_path: str) -> str:
    relative = os.path.relpath(file_path, REPORTES_DIR).replace(os.sep, '/')
    return f'/api/downloads/file?type=report&download={quote(relative, safe="")}'


def _build_open_folder_url(relative_folder: str, filename: Optional[str] = None) -> str:
    query = f'folder={quote(relative_folder, safe="")}'
    if filename:
        query += f'&file={quote(filename, safe="")}'
    return f'/api/reports/open-folder?{query}'


def _collect_eligible_periods() -> List[Dict[str, object]]:
    if not os.path.isdir(DESCARGAS_DIR):
        return []

    period_dirs = []
    for name in os.listdir(DESCARGAS_DIR):
        full_path = os.path.join(DESCARGAS_DIR, name)
        if os.path.isdir(full_path) and re.match(r'^\d{4}$', name):
            period_dirs.append(name)

    period_dirs.sort(reverse=True)

    results: List[Dict[str, object]] = []
    for period in period_dirs:
        report_file = _find_report_for_period(period)
        if not report_file:
            continue

        deducciones_uuids = _extract_deducciones_uuids(report_file)

        folder_counts: Dict[str, int] = {}
        all_have_xml = True
        total_xml = 0

        for folder_name in SOURCE_FOLDERS:
            entries = _collect_xml_entries(period, folder_name)
            raw_count = len(entries)
            if folder_name == 'OTROS_CFDI':
                selected_count = len(_filter_entries_by_allowed_uuids(entries, deducciones_uuids))
            else:
                selected_count = raw_count

            folder_counts[folder_name] = selected_count
            total_xml += selected_count

            if raw_count == 0:
                all_have_xml = False

        if not all_have_xml:
            continue

        results.append(
            {
                'period': period,
                'report_filename': os.path.basename(report_file),
                'xml_counts': folder_counts,
                'total_xml': total_xml,
                'deducciones_uuid_count': len(deducciones_uuids),
            }
        )

    return results


@router.get('/periods')
async def list_eligible_periods():
    periods = _collect_eligible_periods()
    return {
        'success': True,
        'periods': periods,
        'max_zip_bytes': MAX_ZIP_BYTES,
    }


@router.post('/generate')
async def generate_sat_packages(periodo: str = Form('')):
    period = _normalize_period(periodo)
    if not period:
        return JSONResponse(
            status_code=400,
            content={
                'success': False,
                'message': 'Selecciona un ejercicio valido de 4 digitos.',
            },
        )

    eligible_periods = {item['period']: item for item in _collect_eligible_periods()}
    if period not in eligible_periods:
        return JSONResponse(
            status_code=400,
            content={
                'success': False,
                'message': (
                    'El ejercicio seleccionado no es elegible. '
                    'Debe tener XML vigentes en las 3 carpetas y un papel de trabajo del mismo ejercicio.'
                ),
            },
        )

    report_file = _find_report_for_period(period)
    if not report_file:
        return JSONResponse(
            status_code=404,
            content={
                'success': False,
                'message': 'No se encontro el papel de trabajo del ejercicio seleccionado.',
            },
        )

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_relative_folder = f'paquetes_sat/{period}/{timestamp}'
    output_dir = os.path.join(REPORTES_DIR, 'paquetes_sat', period, timestamp)
    os.makedirs(output_dir, exist_ok=True)

    generated_files: List[Dict[str, object]] = []
    warnings: List[str] = []

    try:
        deducciones_uuids = _extract_deducciones_uuids(report_file)

        for folder_name in SOURCE_FOLDERS:
            entries = _collect_xml_entries(period, folder_name)
            if not entries:
                continue

            if folder_name == 'OTROS_CFDI':
                filtered_entries = _filter_entries_by_allowed_uuids(entries, deducciones_uuids)
                if not filtered_entries:
                    warnings.append(
                        'No se genero ZIP de OTROS_CFDI porque no hubo UUID coincidentes '
                        'entre los XML vigentes y la hoja Deducciones del papel de trabajo.'
                    )
                    continue
                entries = filtered_entries

            entry_parts = _split_entries_by_zip_limit(
                entries,
                max_bytes=MAX_ZIP_BYTES,
                temp_dir=output_dir,
                temp_prefix=folder_name.lower(),
            )

            split_required = len(entry_parts) > 1
            for part_index, part_entries in enumerate(entry_parts, start=1):
                suffix = f'_P{part_index}' if split_required else ''
                package_base_name = PACKAGE_NAME_BY_FOLDER.get(folder_name, folder_name)
                zip_name = f'{package_base_name}{suffix}.zip'
                zip_path = os.path.join(output_dir, zip_name)

                _write_zip(zip_path, part_entries)
                zip_size = os.path.getsize(zip_path)
                if zip_size > MAX_ZIP_BYTES:
                    raise ValueError(
                        f'No se pudo crear {zip_name} por debajo de 4 MB. '
                        'Revisa el contenido del ejercicio antes de continuar.'
                    )

                generated_files.append(
                    {
                        'name': zip_name,
                        'category': folder_name,
                        'files_count': len(part_entries),
                        'size_bytes': zip_size,
                        'size_human': _human_size(zip_size),
                        'download_url': _build_download_url(zip_path),
                        'open_url': _build_open_folder_url(output_relative_folder, zip_name),
                    }
                )

        report_name = os.path.basename(report_file)
        report_stem = os.path.splitext(report_name)[0]
        report_zip_name = f'{report_stem}.zip'
        report_zip_path = os.path.join(output_dir, report_zip_name)
        _write_zip(report_zip_path, [(report_file, report_name)])

        report_zip_size = os.path.getsize(report_zip_path)
        generated_files.append(
            {
                'name': report_zip_name,
                'category': 'PAPEL_DE_TRABAJO',
                'files_count': 1,
                'size_bytes': report_zip_size,
                'size_human': _human_size(report_zip_size),
                'download_url': _build_download_url(report_zip_path),
                'open_url': _build_open_folder_url(output_relative_folder, report_zip_name),
            }
        )

        return {
            'success': True,
            'message': 'Paquetes ZIP preparados correctamente para carga en SAT.',
            'period': period,
            'output_folder': output_relative_folder,
            'folder_open_url': _build_open_folder_url(output_relative_folder),
            'packages': generated_files,
            'max_zip_bytes': MAX_ZIP_BYTES,
            'warnings': warnings,
        }

    except ValueError as exc:
        return JSONResponse(
            status_code=400,
            content={
                'success': False,
                'message': str(exc),
            },
        )
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={
                'success': False,
                'message': f'No se pudieron generar los paquetes: {exc}',
            },
        )